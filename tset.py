# OMS 에 수량 포함하게 하는 코드
# code_sec_name = 'RUO3EURR20'
# xpath_code = "//*[text()='" + code_sec_name + "']/parent::td/following::td[8]"
# code_no = 8
# driver.find_element_by_xpath(xpath_code).click()
# time.sleep(1)
# action=ActionChains(driver)
# action.send_keys(code_no).perform()

# 카페 24 자동화 V1.0
# 카페 24에서 주문서를 다운받는다. 송장양식으로 다운 받을것!!!
# 지정된 폴더에 날짜형식으로 저장한다
# 송장 제작
# V 1.0 업데이트 내역
# 송장 자동 변환
# 다음 계획 코드 취합후 OMS 등록
# 필요한 파일 DATA_MAP
# 카페 24 주문 자동화 V2.1 ERP 자동 
# V2.2 추후 버전 업데이트 및 자잘한 버그 수정

# 추후 업데이트 계획
# 업데이트 필요 내용 - 참피온 엑시옴 등록 / 코드 오류시 발생되는 상황 확인
# 대표코드 / 품목코드 / 품목명 / 수량 순으로 데이터 프레임 재 정립
# OMS 등록 순서 동일한 대표코드별 정렬 -> 그후 품목코드 수량순으로 입력


import time
import pandas as pd
import os
import datetime
import xlrd, xlwt
from prettytable import PrettyTable
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from selenium import webdriver
import selenium
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common import action_chains
from selenium.webdriver.common import keys
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

now = datetime.datetime.now()
nowDate = now.strftime('%Y-%m-%d')

wbwt = xlwt.Workbook(encoding='utf-8')
ws = wbwt.add_sheet(nowDate, cell_overwrite_ok=True)

# 카페 24 주문서 경로 설정
Cafe24_file_date = input("날짜를 입력하세요 (ex : 20210412) : ")
Cafe24_file = 'C:/kwakcode/auto_cafe24/주문서/cafe24_' +  str(Cafe24_file_date) + '.xlsx'

#데이터 맵 프레임 워크
date_map_file = 'C:/kwakcode/data_map/date_map_code_v_1_8_6.xlsx'
data_df = pd.read_excel(date_map_file)


Cafe24_df = pd.read_excel(Cafe24_file)

title = "Cafe24 " + nowDate + " 리스트"
hostname = title
table = PrettyTable()
table.title = hostname
table.field_names = ['번호','주문번호','수령인','주문상품명(옵션포함)','수량','상품품목코드']
for i in range(len(Cafe24_df)):
    table.add_row([str(i),str(Cafe24_df.iloc[i,0]),str(Cafe24_df.iloc[i,1]),str(Cafe24_df.iloc[i,9]),str(Cafe24_df.iloc[i,10]),str(Cafe24_df.iloc[i,12])])

print(table)

input("주문확인 대기시간 - 주문서 내역 확인할것")

wb = load_workbook(filename = "C:/kwakcode/auto_cafe24/주문서/송장/송장양식.xlsx")
ws = wb[wb.sheetnames[0]]

row = 2
key1 = f'A{row}'
ws[key1].value = '(온)' + str(Cafe24_df.iloc[0,1]) # 수령인
key2 = f'B{row}'
ws[key2].value = str(Cafe24_df.iloc[0,2]) # 수령인 우편번호
key3 = f'C{row}'
ws[key3].value = str(Cafe24_df.iloc[0,3]) # 수령인 주소
key4 = f'D{row}'
ws[key4].value = str(Cafe24_df.iloc[0,4]) # 수령인 전화번호
key5 = f'E{row}'
ws[key5].value = str(Cafe24_df.iloc[0,5]) # 수령인 휴대전화
key6 = f'F{row}'
ws[key6].value = str(1) # 택배수량
key7 = f'G{row}'
ws[key7].value = str(3) # 선착불
key8 = f'H{row}'
ws[key8].value = str('2,500') # 택배운임
key9 = f'I{row}'
totle_cant = str(Cafe24_df.iloc[0,9]) + ' 수량 : ' + str(Cafe24_df.iloc[0,10]) # 주문상품명 및 수량
ws[key9].value = str(totle_cant)
key10 = f'J{row}'
ws[key10].value = "  "
key11 = f'K{row}'
ws[key11].value = "." + str(Cafe24_df.iloc[0,11])
c = 0
j = 1
row = 3
double_jumun = []
for i in range(len(Cafe24_df['주문번호'])):
    
    if j == len(Cafe24_df['주문번호']) :
        break
    
    if str(Cafe24_df['주문번호'][j]) != str(Cafe24_df['주문번호'][j-1]) :
        double_jumun = []
        key1 = f'A{row}'
        ws[key1].value = '(온)' + str(Cafe24_df.iloc[j,1]) # 수령인
        key2 = f'B{row}'
        ws[key2].value = str(Cafe24_df.iloc[j,2]) # 수령인 우편번호
        key3 = f'C{row}'
        ws[key3].value = str(Cafe24_df.iloc[j,3]) # 수령인 주소
        key4 = f'D{row}'
        ws[key4].value = str(Cafe24_df.iloc[j,4]) # 수령인 전화번호
        key5 = f'E{row}'
        ws[key5].value = str(Cafe24_df.iloc[j,5]) # 수령인 휴대전화
        key6 = f'F{row}'
        ws[key6].value = str(1) # 택배수량
        key7 = f'G{row}'
        ws[key7].value = str(3) # 선착불
        key8 = f'H{row}'
        ws[key8].value = str('2,500') # 택배운임
        
        key9 = f'I{row}'
        totle_cant = str(Cafe24_df.iloc[j,9]) + ' 수량 : ' + str(Cafe24_df.iloc[j,10]) # 주문상품명 및 수량
        ws[key9].value = str(totle_cant)
        
        key10 = f'J{row}'
        ws[key10].value = "  "
        
        key11 = f'K{row}'
        ws[key11].value = "." + str(Cafe24_df.iloc[j,11])
        
        
        c = 0
        
    # 주문번호가 위하고 같으면 발생하는 이벤트
    if str(Cafe24_df['주문번호'][j]) == str(Cafe24_df['주문번호'][j-1]) :
        c = c + 1
        totle_cant_2 = '  /  ' + str(Cafe24_df.iloc[j,9]) + ' 수량 : ' + str(Cafe24_df.iloc[j,10]) # 주문상품명 및 수량
        double_jumun.append(totle_cant_2)
        
        key9 = f'I{row-c}'
        ws[key9].value = str(totle_cant) + str(double_jumun)[2:-2]
    
    j = j + 1    
    row = row + 1
wb.save('C:/kwakcode/auto_cafe24/주문서/송장/Cafe_24_송장_' + nowDate + '.xlsx')

# 빈칸 제거하고 다시 저장하기

Cafe24_file_sosang = 'C:/kwakcode/auto_cafe24/주문서/송장/Cafe_24_송장_' + nowDate + '.xlsx'

Cafe24_sosang_df = pd.read_excel(Cafe24_file_sosang)
cafe_last = Cafe24_sosang_df.dropna()

cafe_last.to_excel(excel_writer='C:/kwakcode/auto_cafe24/주문서/송장/Cafe_24_송장_' + nowDate + '.xlsx',index=None)



# 상품코드기준 수량으로 부분합하기
cafe24_df_sum = Cafe24_df.groupby('상품품목코드')['수량'].sum()
# cafe24_df_sum.index[0] <- 부분합 코드
# cafe24_df_sum[0] <- 부분합이 완료된 수량
print('자동 입력 구간')
print(Cafe24_df.groupby('주문상품명(옵션포함)')['수량'].sum())



options = webdriver.ChromeOptions()
options.add_experimental_option("excludeSwitches", ["enable-logging"])
ti1 = time.time()
driver = webdriver.Chrome("c:/kwakcode/chromedriver",options=options)
driver.implicitly_wait(10)
action = ActionChains(driver)
timeeorl = 1
options = Options()

url = "http://oms.xiomtt.com/"
driver.get(url)
driver.maximize_window ()
driver.find_element_by_id("f_id-inputEl").send_keys("dhkwak",Keys.TAB,"1234")
driver.implicitly_wait(5)
driver.find_element_by_id("image-1009").click()
driver.implicitly_wait(5)
driver.find_element_by_id("expandTool-toolEl").click()
driver.implicitly_wait(5)
driver.find_element_by_xpath("/html/body/div[2]/div[2]/div/div/div/div[3]/div/table/tbody/tr[11]/td/div/span").click() # 국내주문 x path
driver.implicitly_wait(5)
driver.find_element_by_xpath("/html/body/div[2]/div[2]/div/div/div/div[3]/div/table/tbody/tr[11]/td/div/span").click() # 국내주문 x path
driver.implicitly_wait(5)
driver.switch_to_default_content()
driver.switch_to_frame(0)
driver.implicitly_wait(5)
driver.find_element_by_id("grid01Button1").click() # 추가 주문
# 국내주문 선택목록
driver.implicitly_wait(5)
driver.find_element_by_xpath("/html/body/div[7]/div[2]/div/div[2]/span/div/fieldset[1]/div/table/tbody/tr[1]/td[2]/table/tbody/tr/td[2]/table/tbody/tr/td[2]/div").click() # 출고유형
time.sleep(2)
driver.find_element_by_xpath("/html/body/div[9]/div/ul/li[2]").click() # B2C 선택
driver.implicitly_wait(5)
driver.find_element_by_xpath("/html/body/div[7]/div[2]/div/div[2]/span/div/fieldset[2]/div/table/tbody/tr[1]/td[1]/table/tbody/tr/td[2]/table/tbody/tr/td[2]/div").click() # 거래처 선택
driver.implicitly_wait(5)
driver.find_element_by_xpath("/html/body/div[13]/div[2]/div/div[1]/div/span/div/table[2]/tbody/tr/td[2]/div/div/div/table[2]/tbody/tr/td[2]/input").click() # 거래처 
driver.implicitly_wait(5)
driver.find_element_by_xpath("/html/body/div[13]/div[2]/div/div[1]/div/span/div/table[2]/tbody/tr/td[2]/div/div/div/table[2]/tbody/tr/td[2]/input").send_keys("361271")
driver.implicitly_wait(5)
driver.find_element_by_xpath("/html/body/div[13]/div[2]/div/div[2]/div[1]/div/div/div[2]/div/a/span[1]").click() # 검사박스
driver.implicitly_wait(5)
driver.find_element_by_xpath("/html/body/div[13]/div[2]/div/div[2]/div[3]/div/table/tbody/tr/td[1]/div/div").click() # 체크박스
driver.implicitly_wait(5)
driver.find_element_by_xpath("/html/body/div[13]/div[2]/div/div[2]/div[1]/div/div/div[3]/div/a/span[1]").click() # 적용박스
driver.find_element_by_xpath("/html/body/div[7]/div[2]/div/div[2]/span/div/fieldset[4]/div/table/tbody/tr[2]/td[1]/table/tbody/tr/td[2]/textarea").send_keys("Cafe_24 / " + str(nowDate) + "등록건")
# driver.implicitly_wait(5)
# driver.find_elements_by_class_name("x-form-field.x-form-checkbox.x-form-cb")[8].click()# 클래스로 찾아보기 거래처 정보동일 체크 해제
driver.implicitly_wait(5)
# 여기부터 클로우즈 함수임
driver.find_element_by_xpath("/html/body/div[7]/div[2]/div/div[1]/div/div/div[4]/div/a").click() # 저장 버튼
driver.implicitly_wait(5)
driver.find_element_by_xpath("/html/body/div[11]/div[3]/div/div/div[1]/div/a/span[2]").click() # ok 버튼

# 클로우즈 함수 삭제후 이 코딩 복사
time.sleep(1)


# 상품코드기준 수량으로 부분합하기
cafe24_df_sum = Cafe24_df.groupby('상품품목코드')['수량'].sum()
# cafe24_df_sum.index[0] <- 부분합 코드
# cafe24_df_sum[0] <- 부분합이 완료된 수량
#데이터 멥 CAFE -> OMS 코드 전환하는 

def oms_sch_daipo(code_name) : # 대표코드 반환 함수
    for i in range(len(data_df)) :
        if code_name == data_df.iloc[i,3] :
            print("품목명 : " + str(data_df.iloc[i,7]) + " / " + str(data_df.iloc[i,8]) + " / " + str(data_df.iloc[i,9]))
            print("OMS 대표코드(반환되는 값) : " + str(data_df.iloc[i,5]))
            oms_daipo = data_df.iloc[i,5]
    return oms_daipo

def oms_sch_pom(code_name) :
    for i in range(len(data_df)) : # 품목코드 반환 함수
        if code_name == data_df.iloc[i,3] :
            print("품목명 : " + str(data_df.iloc[i,7]) + " / " + str(data_df.iloc[i,8]) + " / " + str(data_df.iloc[i,9]))
            print("OMS 품목코드(반환되는 값) : " + str(data_df.iloc[i,6]))
            oms_pom = data_df.iloc[i,6]
    return oms_pom

def oms_sch_pom_name(code_name) :
    for i in range(len(data_df)) : # 품목이름 반환 함수
        if code_name == data_df.iloc[i,3] :
            print("품목명 : " + str(data_df.iloc[i,7]) + " / " + str(data_df.iloc[i,8]) + " / " + str(data_df.iloc[i,9]))
            oms_pom_name = (str(data_df.iloc[i,7]) + " / " + str(data_df.iloc[i,8]) + " / " + str(data_df.iloc[i,9]))
    return oms_pom_name

def oms_sch_xich_name(code_name) :
    for i in range(len(data_df)) : # 화주사 반환 함수
        if code_name == data_df.iloc[i,3] :
            print("OMS 화주사(반환되는 값) : " + str(data_df.iloc[i,4]))
            oms_pom_xich = data_df.iloc[i,4]            
    return oms_pom_xich

def daipo_click():
    driver.find_element_by_xpath("/html/body/div[3]/div/div[1]/div/div/div[1]/div/div/div[2]/div/a").click() #품목 추가 버튼
    time.sleep(1)
    driver.implicitly_wait(5)
    driver.find_element_by_name("CODE").click() # 대표코드 입력 구간

for i in range(len(cafe24_df_sum)):
    dipath = oms_sch_daipo(str(cafe24_df_sum.index[i]))
    action = ActionChains(driver)
    daipo_click()
    action.send_keys(dipath).perform()
    driver.find_element_by_id("styleGridButton0-btnInnerEl").click()
    time.sleep(2)
    driver.implicitly_wait(5)
    driver.find_elements_by_class_name("x-column-header-text")[68].click() # 조회후 전체 클릭
    time.sleep(2)
    driver.implicitly_wait(5)
    driver.find_element_by_id("styleGridButton1-btnEl").click() # 전체 클릭후 적용
    oms_name = oms_sch_pom(cafe24_df_sum.index[i])
    code_sec_name = oms_name
    xpath_code = "//*[text()='" + code_sec_name + "']/parent::td/following::td[8]" # Xpath 부모자식 건너뛰기 -> OMS 수량 등록 할때 쓰임
    code_no = str(cafe24_df_sum[i])
    time.sleep(1)
    # 코드가 자바 스크립트여서 불러오지 못했을때 스크롤을 내려서 코드를 찾는 반복문
    try :
        driver.implicitly_wait(5)
        driver.find_element_by_xpath(xpath_code).click()
    except :
        while True :
            element = driver.find_element_by_xpath("/html/body/div[3]/div/div[3]/div/div/div/div/div[3]/div")
            element.send_keys(Keys.PAGE_DOWN)
            try :
                driver.implicitly_wait(5)
                driver.find_element_by_xpath(xpath_code).click()
                break
            except :
                pass
                        
    time.sleep(1)
    action=ActionChains(driver)
    action.send_keys(code_no).perform()
    driver.implicitly_wait(5)
    driver.find_element_by_xpath("/html/body/div[3]/div/div[3]/div/div/div/div/div[1]/div/div/div[2]/div/a").click()
    driver.implicitly_wait(5)
    driver.find_element_by_xpath("/html/body/div[11]/div[3]/div/div/div[1]").click()
    driver.implicitly_wait(5)
    time.sleep(1)

