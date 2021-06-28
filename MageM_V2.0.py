# 파일 설명
# 메가엠 발주서를 다운후 인쇄
# 인쇄후 OMS 등록 / 신실장님 구분 / 음성 구분
# 구분후 OMS 등록 및 신실장님 엑셀 리스트 작성 후 음성 리스트 작성

# V1.2 
# 값이 없을때 오류 수정
# 파일명 날짜 입력 받아 적용
# 참피온 엑시옴 구분

# V1.3
# 시간 단축

# V1.4
# XIOM 재고 자동 입력 -> 데이터 맵에 자료가 있어야 함
# 코드를 자동 입력할때 페이지에 안나오는 코드를 못찾는 문제 발생
# v1.5
# 페이지 스크롤후 코드 찾는 기능 추가

# 이후 계획
# 참피온 재고 자동 입력
# 대리점 시트 자동 제작

# V1.7
# 버전 업 오류 수정
# v1.8
# 오류 수정 코드를 못찾았을때 강제적으로 종료되게 설정
# v1.9
# 실장님 및 H-50 탁구대 적색 굵게 표시
# v2.0
# 엑셀 파일 자동 시트 복사하기




import time
from typing import Tuple
import pandas as pd
import datetime
import xlrd, xlwt
from prettytable import PrettyTable
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font


now = datetime.datetime.now()
nowDate = now.strftime('%Y-%m-%d')

wbwt = xlwt.Workbook(encoding='utf-8')
ws = wbwt.add_sheet(nowDate, cell_overwrite_ok=True)
title = "MegaM " + nowDate + " 리스트"

# 메가엠 발주서 경로설정
MegeM_file_date = input("날짜를 입력하세요 (ex : 20210319) : ")
MegaM_file = 'C:/kwakcode/MegaM/메가엠 발주서/' +  str(MegeM_file_date) + ' 메가엠 엑시옴 발주서.xlsx'

#신실장님 파일 저장 경로
sin_mothon_Name = 'C:/kwakcode/MegaM/신실장님 발주서/NEW_신실장님 발주서_신_6월 내역.xlsx'

#음성파일 저장 경로
um_mothon_Name = 'C:/kwakcode/MegaM\음성 발주서/음성 발주서_2021-06월.xlsx'


# 데이터 맵 데이터 프레임 만들기
Data_map_file = 'c:/kwakcode/data_map/date_map_code_v_1_8_8.xlsx'
Data_map_df = pd.read_excel(Data_map_file)

MegaM_df = pd.read_excel(MegaM_file)
hostname = title
table = PrettyTable()
table.title = hostname
table.field_names = ['번호','판매처명','주문번호1','상품명(간단)','주문수량','구매자명','배송지주소']
for i in range(len(MegaM_df)):
    table.add_row([str(i),str(MegaM_df.iloc[i,0]),str(MegaM_df.iloc[i,3]),str(MegaM_df.iloc[i,4]),str(MegaM_df.iloc[i,8]),str(MegaM_df.iloc[i,12]),str(MegaM_df.iloc[i,16])])

print(table)

# 입력구간
# 입력구간의 딜레이를 없애기 위해 한곳에서 시작함
sin_list = None
try :
    sin_list = list(map(int, input('신실장님 리스트 번호( , ) 로 구분 표시 :').split(',')))
except :
    pass

umsung_list = None
try :
    umsung_list = list(map(int, input('음성 대신택배 리스트 번호( , ) 로 구분 표시 :').split(',')))
except :
    pass

xiom_oms_list = None
cham_oms_list = None
try :
    xiom_oms_list = list(map(int, input('Xiom - OMS 출고 리스트 번호( , ) 로 구분 표시 :').split(',')))
except :
    pass
try :
    cham_oms_list = list(map(int, input('Champion - OMS 출고 리스트 번호( , ) 로 구분 표시 :').split(',')))
except :
    pass





#신실장님 리스트 작성 시작

if sin_list != None :
    wb = load_workbook(filename = "C:/kwakcode/MegaM/신실장님 발주서/NEW_신실장님 리스트.xlsx")
    ws = wb[wb.sheetnames[0]]
    ws['D2'] = nowDate
    row = 8
    no_um = 1
    for i in sin_list :
        key1 = f'B{row}'
        ws[key1].value = str(no_um) # 순번
        key2 = f'C{row}'
        if MegaM_df.iloc[i,9] == MegaM_df.iloc[i,12] :
            ws[key2].value = str(MegaM_df.iloc[i,12])
        if MegaM_df.iloc[i,9] != MegaM_df.iloc[i,12] :
            ws[key2].value = str(MegaM_df.iloc[i,9]) + "  /  " + str(MegaM_df.iloc[i,12]) # 주문자 / 받는분
        key3 = f'E{row}'
        ws[key3].value = str(MegaM_df.iloc[i,4]) # 품명
        if MegaM_df.iloc[i,4].find('H-50') == 0 :
            bpr = str(MegaM_df.iloc[i,4])
            bpr_2 = bpr.replace('H-50','BP H-50')
            ws[key3].value = bpr_2
            ws[key3].font = Font(color='ff0000',bold=True)
            
        key4 = f'I{row}'
        ws[key4].value = str(MegaM_df.iloc[i,8]) # 수량
        key5 = f'J{row}'
        ws[key5].value = str(MegaM_df.iloc[i,16]) # 주소
        key6 = f'J{row+1}'
        ws[key6].value = str(MegaM_df.iloc[i,13]) + "  /  " + str(MegaM_df.iloc[i,14]) # 전화번호
        key7 = f'J{row+2}'
        ws[key7].value = str(MegaM_df.iloc[i,12]) # 받는사람
        key8 = f'J{row+3}'
        ws[key8].value = str(MegaM_df.iloc[i,0]) + "  /  " + str(MegaM_df.iloc[i,18]) # 비고
        row = row + 4
        no_um = no_um + 1
    sin_file_Name = 'C:/kwakcode/MegaM/신실장님 발주서/NEW_신실장님 리스트_' + nowDate + '.xlsx'

    wb.save(sin_file_Name)

    from win32com.client import Dispatch

    path1 = sin_file_Name
    path2 = sin_mothon_Name

    xl = Dispatch('Excel.Application')
    xl.Visible = True

    wb1 = xl.Workbooks.Open(Filename=path1)
    wb2 = xl.Workbooks.Open(Filename=path2)

    ws1 = wb1.Worksheets(1)
    ws1.Copy(Before=wb2.Worksheets(1))

    wb2.Close(SaveChanges=True)
    wb1.Close(SaveChanges=True)
    
    xl.Quit()


    # 신실장님 리스트 작성 종료


# 음성 리스트 작성 시작

if umsung_list != None :
    wb_um = load_workbook(filename = "C:/kwakcode/MegaM/음성 발주서/음성발주서.xlsx")
    ws_um = wb_um[wb_um.sheetnames[0]]
    ws_um['D2'] = nowDate
    row = 8
    no_um = 1
    for i in umsung_list :
        key1 = f'B{row}'
        ws_um[key1].value = str(no_um) # 순번
        key2 = f'C{row}'
        if MegaM_df.iloc[i,9] == MegaM_df.iloc[i,12] :
            ws_um[key2].value = str(MegaM_df.iloc[i,12])
        if MegaM_df.iloc[i,9] != MegaM_df.iloc[i,12] :
            ws_um[key2].value = str(MegaM_df.iloc[i,9]) + "  /  " + str(MegaM_df.iloc[i,12]) # 주문자 / 받는분
        key3 = f'E{row}'
        ws_um[key3].value = str(MegaM_df.iloc[i,4]) # 품명
        if MegaM_df.iloc[i,4].find('H-50') == 0 :
            bpr = str(MegaM_df.iloc[i,4])
            bpr_2 = bpr.replace('H-50','BP H-50')
            ws_um[key3].value = bpr_2
            ws_um[key3].font = Font(color='ff0000',bold=True)
        
        key4 = f'I{row}'
        ws_um[key4].value = str(MegaM_df.iloc[i,8]) # 수량
        key5 = f'J{row}'
        ws_um[key5].value = str(MegaM_df.iloc[i,16]) # 주소
        key6 = f'J{row+1}'
        ws_um[key6].value = str(MegaM_df.iloc[i,13]) + "  /  " + str(MegaM_df.iloc[i,14]) # 전화번호
        key7 = f'J{row+2}'
        ws_um[key7].value = str(MegaM_df.iloc[i,12]) # 받는사람
        key8 = f'J{row+3}'
        ws_um[key8].value = str(MegaM_df.iloc[i,0]) + "  /  " + str(MegaM_df.iloc[i,18]) # 비고
        row = row + 4
        no_um = no_um + 1
        
    umsung_file_Name = 'C:/kwakcode/MegaM/음성 발주서/음성 발주서_' + nowDate + '.xlsx'
    
    wb_um.save(umsung_file_Name)
    
    from win32com.client import Dispatch

    path1 = umsung_file_Name
    path2 = um_mothon_Name

    xl = Dispatch('Excel.Application')
    xl.Visible = True

    wb1 = xl.Workbooks.Open(Filename=path1)
    wb2 = xl.Workbooks.Open(Filename=path2)

    ws1 = wb1.Worksheets(1)
    ws1.Copy(Before=wb2.Worksheets(1))

    wb2.Close(SaveChanges=True)
    wb1.Close(SaveChanges=True)

    xl.Quit()



# OMS 등록 시작



if xiom_oms_list or cham_oms_list != None :

    from selenium import webdriver
    import selenium
    from selenium.webdriver.chrome.options import Options
    from selenium.webdriver.common import action_chains
    from selenium.webdriver.common import keys
    from selenium.webdriver.common.keys import Keys
    from selenium.webdriver.common.action_chains import ActionChains
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC
    import sys
    options = webdriver.ChromeOptions()
    options.add_experimental_option("excludeSwitches", ["enable-logging"])
    ti1 = time.time()
    driver = webdriver.Chrome("c:/kwakcode/chromedriver",options=options)
    driver.implicitly_wait(10)
    action = ActionChains(driver)
    timeeorl = 1
    options = Options()
    
    url = "http://oms.xiomtt.com/"
    driver.get(url)
    driver.maximize_window ()
    driver.find_element_by_id("f_id-inputEl").send_keys("dhkwak",Keys.TAB,"1234")
    driver.implicitly_wait(5)
    driver.find_element_by_id("image-1009").click()
    driver.implicitly_wait(5)
    driver.find_element_by_id("expandTool-toolEl").click()
    driver.implicitly_wait(5)
    driver.find_element_by_xpath("/html/body/div[2]/div[2]/div/div/div/div[3]/div/table/tbody/tr[11]/td/div/span").click() # 국내주문 x path
    driver.implicitly_wait(5)
    def open_oms_start_xiom ():
        # 국내 주문 함수 -> 위에 주문까지 취소한다음 적용해야 가능함
        driver.find_element_by_xpath("/html/body/div[2]/div[2]/div/div/div/div[3]/div/table/tbody/tr[11]/td/div/span").click() # 국내주문 x path
        driver.implicitly_wait(5)
        driver.switch_to_default_content()
        driver.switch_to_frame(0)
        driver.implicitly_wait(5)
        driver.find_element_by_id("grid01Button1").click() # 추가 주문
        # 국내주문 선택목록
        driver.implicitly_wait(5)
        driver.find_element_by_xpath("/html/body/div[7]/div[2]/div/div[2]/span/div/fieldset[1]/div/table/tbody/tr[1]/td[2]/table/tbody/tr/td[2]/table/tbody/tr/td[2]/div").click() # 출고유형
        time.sleep(2)
        driver.find_element_by_xpath("/html/body/div[9]/div/ul/li[2]").click() # B2C 선택
        driver.implicitly_wait(5)
        driver.find_element_by_xpath("/html/body/div[7]/div[2]/div/div[2]/span/div/fieldset[2]/div/table/tbody/tr[1]/td[1]/table/tbody/tr/td[2]/table/tbody/tr/td[2]/div").click() # 거래처 선택
        driver.implicitly_wait(5)
        driver.find_element_by_xpath("/html/body/div[13]/div[2]/div/div[1]/div/span/div/table[2]/tbody/tr/td[2]/div/div/div/table[2]/tbody/tr/td[2]/input").click() # 거래처 
        driver.implicitly_wait(5)
        driver.find_element_by_xpath("/html/body/div[13]/div[2]/div/div[1]/div/span/div/table[2]/tbody/tr/td[2]/div/div/div/table[2]/tbody/tr/td[2]/input").send_keys("361271")
        driver.implicitly_wait(5)
        driver.find_element_by_xpath("/html/body/div[13]/div[2]/div/div[2]/div[1]/div/div/div[2]/div/a/span[1]").click() # 검사박스
        driver.implicitly_wait(5)
        driver.find_element_by_xpath("/html/body/div[13]/div[2]/div/div[2]/div[3]/div/table/tbody/tr/td[1]/div/div").click() # 체크박스
        driver.implicitly_wait(5)
        driver.find_element_by_xpath("/html/body/div[13]/div[2]/div/div[2]/div[1]/div/div/div[3]/div/a/span[1]").click() # 적용박스
        driver.implicitly_wait(5)
        driver.find_elements_by_class_name("x-form-field.x-form-checkbox.x-form-cb")[8].click()# 클래스로 찾아보기 거래처 정보동일 체크 해제

    def close_oms_start() :
        # 비고란 쓴 다음 행동
        driver.find_element_by_xpath("/html/body/div[7]/div[2]/div/div[1]/div/div/div[4]/div/a").click()
        driver.implicitly_wait(5)
        driver.find_element_by_xpath("/html/body/div[11]/div[3]/div/div/div[1]/div/a/span[2]").click()
        driver.implicitly_wait(5)
        driver.switch_to_default_content()
        driver.implicitly_wait(5)
        driver.find_element_by_xpath("/html/body/div[4]/div[1]/div[1]/div[2]/div/div/a").click()
        driver.implicitly_wait(5)
    
    def open_oms_start_cham ():
        # 국내 주문 함수 -> 위에 주문까지 취소한다음 적용해야 가능함
        driver.implicitly_wait(5)
        driver.find_element_by_xpath("/html/body/div[2]/div[2]/div/div/div/div[3]/div/table/tbody/tr[11]/td/div/span").click() # 국내주문 x path
        driver.implicitly_wait(5)
        driver.switch_to_default_content()
        driver.implicitly_wait(5)
        driver.switch_to_frame(0)
        driver.implicitly_wait(5)
        driver.find_element_by_id("grid01Button1").click() # 추가 주문
        driver.implicitly_wait(5)
        driver.find_element_by_xpath("/html/body/div[7]/div[2]/div/div[2]/span/div/fieldset[1]/div/table/tbody/tr[1]/td[1]/table/tbody/tr/td[2]/table/tbody/tr/td[2]/div").click() # 참피온 클릭
        time.sleep(2)
        driver.find_element_by_xpath("/html/body/div[9]/div/ul/li[1]").click()
        # 국내주문 선택목록
        driver.implicitly_wait(5)
        driver.find_element_by_xpath("/html/body/div[7]/div[2]/div/div[2]/span/div/fieldset[1]/div/table/tbody/tr[1]/td[2]/table/tbody/tr/td[2]/table/tbody/tr/td[2]/div").click() # 출고유형
        time.sleep(2)
        driver.find_element_by_xpath("/html/body/div[12]/div/ul/li[2]").click() # B2C 선택
        driver.implicitly_wait(5)
        driver.find_element_by_xpath("/html/body/div[7]/div[2]/div/div[2]/span/div/fieldset[2]/div/table/tbody/tr[1]/td[1]/table/tbody/tr/td[2]/table/tbody/tr/td[2]/div").click() # 거래처 선택
        driver.implicitly_wait(5)
        driver.find_element_by_xpath("/html/body/div[16]/div[2]/div/div[1]/div/span/div/table[2]/tbody/tr/td[2]/div/div/div/table[2]/tbody/tr/td[2]/input").click() # 거래처
        driver.implicitly_wait(5)
        driver.find_element_by_xpath("/html/body/div[16]/div[2]/div/div[1]/div/span/div/table[2]/tbody/tr/td[2]/div/div/div/table[2]/tbody/tr/td[2]/input").send_keys("361271")
        driver.implicitly_wait(5)
        driver.find_element_by_xpath("/html/body/div[16]/div[2]/div/div[2]/div[1]/div/div/div[2]/div/a/span[1]").click() # 검사박스
        driver.implicitly_wait(5)
        driver.find_element_by_xpath("/html/body/div[16]/div[2]/div/div[2]/div[3]/div/table/tbody/tr/td[1]/div/div").click() # 체크박스
        driver.implicitly_wait(5)
        driver.find_element_by_xpath("/html/body/div[16]/div[2]/div/div[2]/div[1]/div/div/div[3]/div/a/span[1]").click() # 적용박스
        driver.implicitly_wait(5)
        driver.find_elements_by_class_name("x-form-field.x-form-checkbox.x-form-cb")[8].click()# 클래스로 찾아보기 거래처 정보동일 체크 해제

    if xiom_oms_list != None :
        Daipo_code = []
        Pummok_code = []
        for i in xiom_oms_list :
            print('메가엠 주문자 : ' + MegaM_df.iloc[i,9]) # MegaM 상품명
            print('메가엠 품명 : ' + MegaM_df.iloc[i,4]) # MegaM 상품명
            MegaM_sch = str(MegaM_df.iloc[i,5]) + str(MegaM_df.iloc[i,6]) + str(MegaM_df.iloc[i,7]) # MegaM 코드 1+2+3
            for j in range(len(Data_map_df)) :
                Data_map_sch = str(Data_map_df.iloc[j,14]) + str(Data_map_df.iloc[j,15]) + str(Data_map_df.iloc[j,16])
                if MegaM_sch == Data_map_sch :
                    print('OMS 대표코드 : ' + Data_map_df.iloc[j,5]) # OMS 대표코드
                    Daipo_code.append(str(Data_map_df.iloc[j,5]))
                    print('OMS 폼목코드 : ' + Data_map_df.iloc[j,6]) # OMS 품목코드
                    Pummok_code.append(str(Data_map_df.iloc[j,6]))
                    print('OMS 품목명 : ' + Data_map_df.iloc[j,7])
                    print('OMS 규격1 : ' + str(Data_map_df.iloc[j,8]))
                    print('OMS 규격2 : ' + str(Data_map_df.iloc[j,9]))
        if len(xiom_oms_list) != len(Daipo_code):
            sys.exit()

        k = 0
        for i in xiom_oms_list:
            open_oms_start_xiom()
            driver.implicitly_wait(5)
            driver.find_element_by_xpath("/html/body/div[7]/div[2]/div/div[2]/span/div/fieldset[3]/div/table/tbody/tr[2]/td[1]/table/tbody/tr/td[2]/input").send_keys(str(MegaM_df.iloc[i,12]))
            driver.implicitly_wait(5)
            driver.find_element_by_xpath("/html/body/div[7]/div[2]/div/div[2]/span/div/fieldset[3]/div/table/tbody/tr[2]/td[2]/table/tbody/tr/td[2]/input").send_keys(str(MegaM_df.iloc[i,13]))
            driver.implicitly_wait(5)
            driver.find_element_by_xpath("/html/body/div[7]/div[2]/div/div[2]/span/div/fieldset[3]/div/table/tbody/tr[3]/td/table/tbody/tr/td[2]/input").send_keys(str(MegaM_df.iloc[i,16]))
            driver.implicitly_wait(5)
            driver.find_element_by_xpath("/html/body/div[7]/div[2]/div/div[2]/span/div/fieldset[4]/div/table/tbody/tr[2]/td[1]/table/tbody/tr/td[2]/textarea").send_keys("메가엠 / " + str(MegaM_df.iloc[i,12]))
            driver.implicitly_wait(5)
            driver.implicitly_wait(5)
            # 여기부터 클로우즈 함수임
            driver.find_element_by_xpath("/html/body/div[7]/div[2]/div/div[1]/div/div/div[4]/div/a").click() # 저장 버튼
            driver.implicitly_wait(5)
            driver.find_element_by_xpath("/html/body/div[11]/div[3]/div/div/div[1]/div/a/span[2]").click() # ok 버튼
            driver.implicitly_wait(5)
            # 클로우즈 함수 삭제후 이 코딩 복사
            time.sleep(3)
            driver.find_element_by_xpath("/html/body/div[3]/div/div[1]/div/div/div[1]/div/div/div[2]/div/a").click() #품목 추가 버튼
            time.sleep(3)
            action = ActionChains(driver)
            driver.find_element_by_name("CODE").send_keys(str(Daipo_code[k]))
            # driver.find_element_by_xpath("/html/body/div[16]/div[2]/div/div/div[1]/div/span/div/table[2]/tbody/tr/td[2]/div/div/div/table[2]/tbody/tr/td[2]/input").send_keys(str(Daipo_code[k]))
            driver.implicitly_wait(5)
            driver.find_element_by_xpath("/html/body/div[16]/div[2]/div/div/div[2]/div[1]/div/div/div[2]/div/a").click() # 조회버튼
            driver.implicitly_wait(5)
            driver.find_element_by_xpath("/html/body/div[16]/div[2]/div/div/div[2]/div[2]/div/div/div[1]/div").click() # 전체 클릭
            driver.implicitly_wait(5)
            driver.find_element_by_xpath("/html/body/div[16]/div[2]/div/div/div[2]/div[1]/div/div/div[3]/div/a").click()
            time.sleep(1)
            code_sec_name = str(Pummok_code[k])
            xpath_code = "//*[text()='" + code_sec_name + "']/parent::td/following::td[8]" # Xpath 부모자식 건너뛰기 -> OMS 수량 등록 할때 쓰임
            code_no = str(MegaM_df.iloc[i,8])
            time.sleep(1)
            # 코드가 자바 스크립트여서 불러오지 못했을때 스크롤을 내려서 코드를 찾는 반복문
            try :
                driver.implicitly_wait(5)
                driver.find_element_by_xpath(xpath_code).click()
            except :
                while True :
                    element = driver.find_element_by_xpath("/html/body/div[3]/div/div[3]/div/div/div/div/div[3]/div")
                    element.send_keys(Keys.PAGE_DOWN)
                    try :
                        driver.implicitly_wait(5)
                        driver.find_element_by_xpath(xpath_code).click()
                        break
                    except :
                        pass
            
            time.sleep(1)
            action=ActionChains(driver)
            action.send_keys(code_no).perform()
            driver.implicitly_wait(5)
            driver.find_element_by_xpath("/html/body/div[3]/div/div[3]/div/div/div/div/div[1]/div/div/div[2]/div/a").click()
            driver.implicitly_wait(5)
            driver.find_element_by_xpath("/html/body/div[11]/div[3]/div/div/div[1]").click()
            driver.implicitly_wait(5)
            driver.switch_to_default_content()
            driver.implicitly_wait(5)
            driver.find_element_by_xpath("/html/body/div[4]/div[1]/div[1]/div[2]/div/div/a").click()
            time.sleep(1)
            k = k + 1

    if cham_oms_list != None :
        for i in cham_oms_list:
            open_oms_start_cham()
            driver.implicitly_wait(5)
            driver.find_element_by_xpath("/html/body/div[7]/div[2]/div/div[2]/span/div/fieldset[3]/div/table/tbody/tr[2]/td[1]/table/tbody/tr/td[2]/input").send_keys(str(MegaM_df.iloc[i,12]))
            driver.implicitly_wait(5)
            driver.find_element_by_xpath("/html/body/div[7]/div[2]/div/div[2]/span/div/fieldset[3]/div/table/tbody/tr[2]/td[2]/table/tbody/tr/td[2]/input").send_keys(str(MegaM_df.iloc[i,13]))
            driver.implicitly_wait(5)
            driver.find_element_by_xpath("/html/body/div[7]/div[2]/div/div[2]/span/div/fieldset[3]/div/table/tbody/tr[3]/td/table/tbody/tr/td[2]/input").send_keys(str(MegaM_df.iloc[i,16]))
            driver.implicitly_wait(5)
            driver.find_element_by_xpath("/html/body/div[7]/div[2]/div/div[2]/span/div/fieldset[4]/div/table/tbody/tr[2]/td[1]/table/tbody/tr/td[2]/textarea").send_keys("메가엠 / " + str(MegaM_df.iloc[i,12]))
            driver.implicitly_wait(5)
            close_oms_start()

